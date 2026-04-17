"""Excel (.xlsx) file handler using openpyxl."""

from __future__ import annotations

import io

from openpyxl import load_workbook

from entity_registry import EntityRegistry


def read_excel(file_bytes: bytes) -> tuple[object, str]:
    """Return (workbook, full_text_for_detection)."""
    wb = load_workbook(io.BytesIO(file_bytes))
    texts: list[str] = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    texts.append(str(cell.value))
    return wb, "\n".join(texts)


def write_excel(
    wb,
    registry: EntityRegistry,
    skip_entities: set[str] | None = None,
) -> bytes:
    """Replace all mapped entities in every cell and return xlsx bytes."""
    if skip_entities is None:
        skip_entities = set()

    mapping = {
        row["original"]: row["pseudonym"]
        for row in registry.mapping_table
        if row["original"] not in skip_entities
    }

    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None and isinstance(cell.value, str):
                    val = cell.value
                    for orig, pseudo in mapping.items():
                        val = val.replace(orig, pseudo)
                    cell.value = val

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
